"""
Excel Export Service for LMS Analytics
Generates Excel files with detailed analytics and charts
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from typing import List, Dict, Any, Optional
from datetime import datetime
from io import BytesIO


class ExcelExportService:
    """Service for exporting analytics data to Excel with charts"""
    
    # Color scheme
    COLORS = {
        'header': 'FF4472C4',
        'excellent': 'FF92D050',  # Green
        'good': 'FFFFFF00',        # Yellow
        'needs_attention': 'FFFF0000',  # Red
        'border': 'FF000000'
    }
    
    def create_analytics_workbook(
        self,
        course_name: str,
        students_data: List[Dict[str, Any]],
        course_overview: Optional[Dict[str, Any]] = None,
        groups_data: Optional[List[Dict[str, Any]]] = None
    ) -> BytesIO:
        """
        Create Excel workbook with analytics data and charts
        
        Args:
            course_name: Name of the course
            students_data: List of student analytics dictionaries
            course_overview: Course overview statistics
            groups_data: List of group analytics dictionaries
            
        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_student_progress_sheet(wb, students_data, course_name)
        
        if course_overview:
            self._create_course_overview_sheet(wb, course_overview, course_name)
        
        if groups_data and len(groups_data) > 0:
            self._create_groups_summary_sheet(wb, groups_data)
        
        # Create charts sheet
        if students_data:
            self._create_charts_sheet(wb, students_data, groups_data)
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_student_progress_sheet(
        self,
        wb: Workbook,
        students_data: List[Dict[str, Any]],
        course_name: str
    ):
        """Create Student Progress sheet with detailed metrics"""
        ws = wb.create_sheet("Student Progress")
        
        # Title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"Student Progress - {course_name}"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                     end_color=self.COLORS['header'], 
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Headers
        headers = [
            'Student Name', 'Email', 'Student ID', 'Groups', 
            'Progress %', 'Completed Steps', 'Total Steps',
            'Assignments Done', 'Total Assignments', 'Score %',
            'Study Time (min)', 'Streak (days)', 'Last Activity', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                   end_color=self.COLORS['header'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[2].height = 25
        
        # Data rows
        for idx, student in enumerate(students_data, 3):
            progress_pct = student.get('progress_percentage', 0)
            score_pct = student.get('average_score', 0)
            
            row_data = [
                student.get('student_name', 'N/A'),
                student.get('email', 'N/A'),
                student.get('student_id', 'N/A'),
                ', '.join(student.get('groups', [])) if student.get('groups') else 'No groups',
                progress_pct,
                student.get('completed_steps', 0),
                student.get('total_steps', 0),
                student.get('assignments_completed', 0),
                student.get('total_assignments', 0),
                score_pct,
                student.get('total_study_time', 0),
                student.get('current_streak', 0),
                student.get('last_activity', 'Never'),
                self._get_status(progress_pct)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Conditional formatting for progress
                if col == 5:  # Progress %
                    if progress_pct >= 80:
                        cell.fill = PatternFill(start_color=self.COLORS['excellent'], 
                                              end_color=self.COLORS['excellent'], 
                                              fill_type='solid')
                    elif progress_pct >= 50:
                        cell.fill = PatternFill(start_color=self.COLORS['good'], 
                                              end_color=self.COLORS['good'], 
                                              fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color=self.COLORS['needs_attention'], 
                                              end_color=self.COLORS['needs_attention'], 
                                              fill_type='solid')
                
                # Format percentages
                if col in [5, 10]:
                    cell.number_format = '0.00"%"'
        
        # Freeze panes
        ws.freeze_panes = 'A3'
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Widen some columns
        ws.column_dimensions['A'].width = 25  # Student Name
        ws.column_dimensions['B'].width = 30  # Email
        ws.column_dimensions['D'].width = 30  # Groups
    
    def _create_course_overview_sheet(
        self,
        wb: Workbook,
        course_overview: Dict[str, Any],
        course_name: str
    ):
        """Create Course Overview sheet"""
        ws = wb.create_sheet("Course Overview")
        
        # Title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = f"Course Overview - {course_name}"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                     end_color=self.COLORS['header'], 
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Overview data
        overview_data = [
            ('Course Name', course_overview.get('course_name', 'N/A')),
            ('Total Students', course_overview.get('total_students', 0)),
            ('Average Progress', f"{course_overview.get('average_progress', 0):.2f}%"),
            ('', ''),
            ('Course Structure', ''),
            ('Total Modules', course_overview.get('total_modules', 0)),
            ('Total Lessons', course_overview.get('total_lessons', 0)),
            ('Total Steps', course_overview.get('total_steps', 0)),
            ('Total Assignments', course_overview.get('total_assignments', 0)),
            ('', ''),
            ('Engagement Metrics', ''),
            ('Active Students (>0% progress)', course_overview.get('active_students', 0)),
            ('Students with >50% progress', course_overview.get('students_above_50', 0)),
            ('Students with >80% progress', course_overview.get('students_above_80', 0)),
            ('Average Study Time (min)', f"{course_overview.get('average_study_time', 0):.1f}"),
            ('', ''),
            ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for idx, (label, value) in enumerate(overview_data, 3):
            label_cell = ws.cell(row=idx, column=1, value=label)
            value_cell = ws.cell(row=idx, column=2, value=value)
            
            if label and not value:  # Section headers
                label_cell.font = Font(bold=True, size=12)
                ws.merge_cells(f'A{idx}:B{idx}')
            else:
                label_cell.font = Font(bold=True)
                label_cell.alignment = Alignment(horizontal='right')
                value_cell.alignment = Alignment(horizontal='left')
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
    
    def _create_groups_summary_sheet(
        self,
        wb: Workbook,
        groups_data: List[Dict[str, Any]]
    ):
        """Create Groups Summary sheet"""
        ws = wb.create_sheet("Groups Summary")
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Groups Summary"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                     end_color=self.COLORS['header'], 
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Headers
        headers = ['Group Name', 'Students', 'Avg Progress %', 'Teacher/Curator', 'Active Students', 'Completion Rate']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                   end_color=self.COLORS['header'], 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[2].height = 25
        
        # Data rows
        for idx, group in enumerate(groups_data, 3):
            avg_progress = group.get('average_progress', 0)
            
            row_data = [
                group.get('group_name', 'N/A'),
                group.get('student_count', 0),
                avg_progress,
                group.get('teacher_name', 'N/A') or group.get('curator_name', 'N/A'),
                group.get('active_students', 0),
                f"{(group.get('active_students', 0) / max(group.get('student_count', 1), 1) * 100):.1f}%"
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if col == 3:  # Progress %
                    cell.number_format = '0.00"%"'
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        ws.freeze_panes = 'A3'
    
    def _create_charts_sheet(
        self,
        wb: Workbook,
        students_data: List[Dict[str, Any]],
        groups_data: Optional[List[Dict[str, Any]]] = None
    ):
        """Create Charts sheet with visualizations"""
        ws = wb.create_sheet("Charts & Analytics")
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "Analytics Charts"
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                     end_color=self.COLORS['header'], 
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        current_row = 3
        
        # Chart 1: Progress Distribution
        current_row = self._create_progress_distribution_chart(ws, students_data, current_row)
        
        # Chart 2: Top 10 Students
        current_row = self._create_top_students_chart(ws, students_data, current_row + 2)
        
        # Chart 3: Groups Comparison (if groups data available)
        if groups_data and len(groups_data) > 0:
            current_row = self._create_groups_comparison_chart(ws, groups_data, current_row + 2)
    
    def _create_progress_distribution_chart(
        self,
        ws,
        students_data: List[Dict[str, Any]],
        start_row: int
    ) -> int:
        """Create progress distribution bar chart"""
        # Calculate distribution
        bins = {'0-49%': 0, '50-79%': 0, '80-100%': 0}
        
        for student in students_data:
            progress = student.get('progress_percentage', 0)
            if progress < 50:
                bins['0-49%'] += 1
            elif progress < 80:
                bins['50-79%'] += 1
            else:
                bins['80-100%'] += 1
        
        # Write data
        ws.cell(row=start_row, column=1, value="Progress Range").font = Font(bold=True)
        ws.cell(row=start_row, column=2, value="Number of Students").font = Font(bold=True)
        
        data_start = start_row + 1
        for idx, (range_label, count) in enumerate(bins.items(), data_start):
            ws.cell(row=idx, column=1, value=range_label)
            ws.cell(row=idx, column=2, value=count)
        
        # Create chart
        chart = BarChart()
        chart.title = "Student Progress Distribution"
        chart.x_axis.title = "Progress Range"
        chart.y_axis.title = "Number of Students"
        
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 3)
        categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 3)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 10
        chart.width = 20
        
        ws.add_chart(chart, f'D{start_row}')
        
        return start_row + 3
    
    def _create_top_students_chart(
        self,
        ws,
        students_data: List[Dict[str, Any]],
        start_row: int
    ) -> int:
        """Create top 10 students bar chart"""
        # Sort by progress
        sorted_students = sorted(
            students_data,
            key=lambda x: x.get('progress_percentage', 0),
            reverse=True
        )[:10]
        
        # Write data
        ws.cell(row=start_row, column=1, value="Student").font = Font(bold=True)
        ws.cell(row=start_row, column=2, value="Progress %").font = Font(bold=True)
        
        data_start = start_row + 1
        for idx, student in enumerate(sorted_students, data_start):
            ws.cell(row=idx, column=1, value=student.get('student_name', 'N/A'))
            ws.cell(row=idx, column=2, value=student.get('progress_percentage', 0))
        
        # Create chart
        chart = BarChart()
        chart.title = "Top 10 Students by Progress"
        chart.x_axis.title = "Student"
        chart.y_axis.title = "Progress %"
        chart.type = "col"  # Vertical bars
        
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(sorted_students))
        categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(sorted_students))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 12
        chart.width = 20
        
        ws.add_chart(chart, f'D{start_row}')
        
        return start_row + len(sorted_students)
    
    def _create_groups_comparison_chart(
        self,
        ws,
        groups_data: List[Dict[str, Any]],
        start_row: int
    ) -> int:
        """Create groups comparison bar chart"""
        # Write data
        ws.cell(row=start_row, column=1, value="Group").font = Font(bold=True)
        ws.cell(row=start_row, column=2, value="Avg Progress %").font = Font(bold=True)
        
        data_start = start_row + 1
        for idx, group in enumerate(groups_data, data_start):
            ws.cell(row=idx, column=1, value=group.get('group_name', 'N/A'))
            ws.cell(row=idx, column=2, value=group.get('average_progress', 0))
        
        # Create chart
        chart = BarChart()
        chart.title = "Average Progress by Group"
        chart.x_axis.title = "Group"
        chart.y_axis.title = "Average Progress %"
        chart.type = "col"
        
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(groups_data))
        categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(groups_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 12
        chart.width = 20
        
        ws.add_chart(chart, f'D{start_row}')
        
        return start_row + len(groups_data)
    
    def _get_status(self, progress: float) -> str:
        """Get status label based on progress"""
        if progress >= 80:
            return "Excellent"
        elif progress >= 50:
            return "Good"
        elif progress > 0:
            return "Needs Attention"
        else:
            return "Not Started"


# Singleton instance
_export_service = None

def get_excel_export_service() -> ExcelExportService:
    """Get or create Excel export service instance"""
    global _export_service
    
    if _export_service is None:
        _export_service = ExcelExportService()
    
    return _export_service
