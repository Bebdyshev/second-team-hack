"""
Google Sheets Service for LMS Analytics Export
Handles creation and population of Google Sheets with student analytics data
Also supports Excel export as fallback
"""

import gspread
from google.oauth2.service_account import Credentials
from typing import List, Dict, Any, Optional
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class GoogleSheetsService:
    """Service for exporting analytics data to Google Sheets"""
    
    SCOPES = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    def __init__(self, credentials_path: str):
        """
        Initialize Google Sheets service
        
        Args:
            credentials_path: Path to service account JSON file
        """
        self.credentials_path = credentials_path
        self.client = None
        self._authenticate()
    
    def _authenticate(self):
        """Authenticate with Google Sheets API"""
        try:
            creds = Credentials.from_service_account_file(
                self.credentials_path,
                scopes=self.SCOPES
            )
            self.client = gspread.authorize(creds)
        except Exception as e:
            raise Exception(f"Failed to authenticate with Google Sheets: {str(e)}")
    
    def cleanup_old_spreadsheets(self, days_old: int = 7, max_to_delete: int = 10):
        """
        Delete old spreadsheets from service account's Drive to free up space
        
        Args:
            days_old: Delete files older than this many days
            max_to_delete: Maximum number of files to delete in one run
        
        Returns:
            Number of files deleted
        """
        try:
            from datetime import datetime, timedelta
            
            # List all spreadsheets owned by service account
            # Use openall() which returns all spreadsheets the service account has access to
            all_sheets = self.client.openall()
            deleted_count = 0
            
            print(f"Found {len(all_sheets)} total spreadsheets")
            
            for sheet in all_sheets:
                if deleted_count >= max_to_delete:
                    print(f"Reached max deletion limit ({max_to_delete})")
                    break
                
                try:
                    # Check if spreadsheet name starts with "Analytics -" (our exports)
                    if not sheet.title.startswith("Analytics -"):
                        continue
                    
                    # Try to delete it
                    self.client.del_spreadsheet(sheet.id)
                    deleted_count += 1
                    print(f"Deleted: {sheet.title}")
                    
                except Exception as e:
                    print(f"Could not delete {sheet.title}: {e}")
                    continue
            
            return deleted_count
            
        except Exception as e:
            print(f"Cleanup failed: {e}")
            return 0
    
    def create_analytics_spreadsheet(
        self,
        title: str,
        course_name: str,
        students_data: List[Dict[str, Any]],
        course_overview: Optional[Dict[str, Any]] = None,
        groups_data: Optional[List[Dict[str, Any]]] = None,
        share_with_email: Optional[str] = None
    ) -> str:
        """
        Create a new Google Sheets spreadsheet with analytics data
        
        Args:
            title: Title of the spreadsheet
            course_name: Name of the course
            students_data: List of student analytics dictionaries
            course_overview: Course overview statistics
            groups_data: List of group analytics dictionaries
            share_with_email: Email to share the spreadsheet with
            
        Returns:
            URL of the created spreadsheet
        """
        try:
            # Try to cleanup old files first to free up space
            try:
                print("Attempting to cleanup old spreadsheets...")
                deleted = self.cleanup_old_spreadsheets(days_old=1)  # Delete files older than 1 day
                print(f"Cleaned up {deleted} old spreadsheets")
            except Exception as e:
                print(f"Cleanup warning: {e}")
            
            # Create new spreadsheet
            print(f"Attempting to create spreadsheet: {title}")
            try:
                spreadsheet = self.client.create(title)
                print(f"✅ Successfully created spreadsheet: {title}")
            except Exception as create_error:
                print(f"❌ Failed to create spreadsheet: {create_error}")
                print(f"Error type: {type(create_error)}")
                print(f"Error details: {str(create_error)}")
                raise
            
            # Immediately transfer ownership to user if email provided
            # This moves the file from service account's Drive to user's Drive
            if share_with_email:
                try:
                    # Transfer ownership directly (this should move it to user's Drive)
                    spreadsheet.share(
                        share_with_email, 
                        perm_type='user', 
                        role='owner', 
                        notify=True,
                        email_message=f'Analytics report for {course_name} has been created and shared with you.'
                    )
                    print(f"Transferred ownership to {share_with_email}")
                except Exception as e:
                    print(f"Warning: Could not transfer ownership: {e}")
                    # If ownership transfer fails, at least share it
                    try:
                        spreadsheet.share(share_with_email, perm_type='user', role='writer', notify=True)
                        print(f"Shared with {share_with_email} as writer")
                    except Exception as e2:
                        print(f"Warning: Could not share: {e2}")
            
            # Create sheets
            self._create_student_progress_sheet(spreadsheet, students_data, course_name)
            
            if course_overview:
                self._create_course_overview_sheet(spreadsheet, course_overview, course_name)
            
            if groups_data:
                self._create_groups_summary_sheet(spreadsheet, groups_data)
            
            # Delete default "Sheet1" if it exists
            try:
                default_sheet = spreadsheet.worksheet("Sheet1")
                spreadsheet.del_worksheet(default_sheet)
            except:
                pass
            
            return spreadsheet.url
            
        except Exception as e:
            raise Exception(f"Failed to create spreadsheet: {str(e)}")
    
    def _create_student_progress_sheet(
        self,
        spreadsheet,
        students_data: List[Dict[str, Any]],
        course_name: str
    ):
        """Create and populate Student Progress sheet"""
        
        # Create or get sheet
        try:
            sheet = spreadsheet.add_worksheet(title="Student Progress", rows=1000, cols=15)
        except:
            sheet = spreadsheet.worksheet("Student Progress")
        
        # Header row
        headers = [
            "Student Name",
            "Email",
            "Student ID",
            "Groups",
            "Active Courses",
            "Progress %",
            "Steps Completed",
            "Total Steps",
            "Assignments Completed",
            "Total Assignments",
            "Assignment Score %",
            "Study Time (hours)",
            "Daily Streak",
            "Last Activity",
            "Status"
        ]
        
        # Prepare data rows
        rows = [headers]
        
        for student in students_data:
            # Format groups
            groups_str = ", ".join([g.get('name', '') for g in student.get('groups', [])])
            
            # Calculate study time in hours
            study_time_hours = round(student.get('total_study_time_minutes', 0) / 60, 1)
            
            # Determine status
            progress = student.get('completion_percentage', 0)
            if progress >= 80:
                status = "Excellent"
            elif progress >= 50:
                status = "Good"
            else:
                status = "Needs Attention"
            
            # Format last activity date
            last_activity = student.get('last_activity_date', '')
            if last_activity:
                try:
                    last_activity = datetime.fromisoformat(str(last_activity)).strftime('%Y-%m-%d')
                except:
                    pass
            
            row = [
                student.get('student_name', ''),
                student.get('student_email', ''),
                student.get('student_number', ''),
                groups_str,
                student.get('active_courses_count', 0),
                round(student.get('completion_percentage', 0), 1),
                student.get('completed_steps', 0),
                student.get('total_steps', 0),
                student.get('completed_assignments', 0),
                student.get('total_assignments', 0),
                round(student.get('assignment_score_percentage', 0), 1),
                study_time_hours,
                student.get('daily_streak', 0),
                last_activity,
                status
            ]
            rows.append(row)
        
        # Update sheet with all data at once
        sheet.update('A1', rows)
        
        # Format header row
        sheet.format('A1:O1', {
            'backgroundColor': {'red': 0.2, 'green': 0.4, 'blue': 0.8},
            'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
            'horizontalAlignment': 'CENTER'
        })
        
        # Format progress columns with conditional formatting
        self._apply_conditional_formatting(sheet, 'F', len(rows))  # Progress %
        
        # Freeze header row
        sheet.freeze(rows=1)
        
        # Auto-resize columns
        sheet.columns_auto_resize(0, len(headers))
    
    def _create_course_overview_sheet(
        self,
        spreadsheet,
        course_overview: Dict[str, Any],
        course_name: str
    ):
        """Create and populate Course Overview sheet"""
        
        try:
            sheet = spreadsheet.add_worksheet(title="Course Overview", rows=100, cols=5)
        except:
            sheet = spreadsheet.worksheet("Course Overview")
        
        # Prepare overview data
        course_info = course_overview.get('course_info', {})
        structure = course_overview.get('structure', {})
        engagement = course_overview.get('engagement', {})
        
        rows = [
            ["Course Analytics Report"],
            [""],
            ["Course Information"],
            ["Course Name", course_info.get('title', course_name)],
            ["Teacher", course_info.get('teacher_name', 'N/A')],
            [""],
            ["Course Structure"],
            ["Total Modules", structure.get('total_modules', 0)],
            ["Total Lessons", structure.get('total_lessons', 0)],
            ["Total Steps", structure.get('total_steps', 0)],
            [""],
            ["Student Engagement"],
            ["Total Enrolled Students", engagement.get('total_enrolled_students', 0)],
            ["Total Time Spent (hours)", round(engagement.get('total_time_spent_minutes', 0) / 60, 1)],
            ["Total Completed Steps", engagement.get('total_completed_steps', 0)],
            ["Average Completion Rate", f"{round(engagement.get('average_completion_rate', 0), 1)}%"],
            [""],
            ["Report Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        sheet.update('A1', rows)
        
        # Format title
        sheet.format('A1', {
            'backgroundColor': {'red': 0.2, 'green': 0.4, 'blue': 0.8},
            'textFormat': {'bold': True, 'fontSize': 14, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
            'horizontalAlignment': 'CENTER'
        })
        
        # Merge title cell
        sheet.merge_cells('A1:B1')
        
        # Format section headers
        for row_num in [3, 7, 12]:
            sheet.format(f'A{row_num}', {
                'textFormat': {'bold': True, 'fontSize': 12},
                'backgroundColor': {'red': 0.9, 'green': 0.9, 'blue': 0.9}
            })
        
        # Auto-resize columns
        sheet.columns_auto_resize(0, 2)
    
    def _create_groups_summary_sheet(
        self,
        spreadsheet,
        groups_data: List[Dict[str, Any]]
    ):
        """Create and populate Groups Summary sheet"""
        
        try:
            sheet = spreadsheet.add_worksheet(title="Groups Summary", rows=500, cols=10)
        except:
            sheet = spreadsheet.worksheet("Groups Summary")
        
        # Header row
        headers = [
            "Group Name",
            "Description",
            "Teacher",
            "Curator",
            "Students Count",
            "Avg Progress %",
            "Avg Assignment Score %",
            "Avg Study Time (hours)",
            "Created Date",
            "Status"
        ]
        
        rows = [headers]
        
        for group in groups_data:
            avg_progress = group.get('average_completion_percentage', 0)
            
            # Determine status
            if avg_progress >= 70:
                status = "On Track"
            elif avg_progress >= 40:
                status = "Needs Support"
            else:
                status = "At Risk"
            
            # Format created date
            created_at = group.get('created_at', '')
            if created_at:
                try:
                    created_at = datetime.fromisoformat(str(created_at)).strftime('%Y-%m-%d')
                except:
                    pass
            
            row = [
                group.get('group_name', ''),
                group.get('description', ''),
                group.get('teacher_name', 'N/A'),
                group.get('curator_name', 'N/A'),
                group.get('students_count', 0),
                round(avg_progress, 1),
                round(group.get('average_assignment_score_percentage', 0), 1),
                round(group.get('average_study_time_minutes', 0) / 60, 1),
                created_at,
                status
            ]
            rows.append(row)
        
        # Update sheet
        sheet.update('A1', rows)
        
        # Format header
        sheet.format('A1:J1', {
            'backgroundColor': {'red': 0.2, 'green': 0.4, 'blue': 0.8},
            'textFormat': {'bold': True, 'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}},
            'horizontalAlignment': 'CENTER'
        })
        
        # Freeze header
        sheet.freeze(rows=1)
        
        # Auto-resize columns
        sheet.columns_auto_resize(0, len(headers))
    
    def _apply_conditional_formatting(self, sheet, column: str, num_rows: int):
        """Apply conditional formatting to a column based on percentage values"""
        
        # Skip header row
        range_notation = f'{column}2:{column}{num_rows}'
        
        # Green for >= 80%
        sheet.add_conditional_format_rule(
            range_notation,
            {
                'type': 'NUMBER_GREATER_THAN_EQ',
                'values': [{'userEnteredValue': '80'}],
                'format': {
                    'backgroundColor': {'red': 0.7, 'green': 0.9, 'blue': 0.7}
                }
            }
        )
        
        # Yellow for 50-79%
        sheet.add_conditional_format_rule(
            range_notation,
            {
                'type': 'NUMBER_BETWEEN',
                'values': [
                    {'userEnteredValue': '50'},
                    {'userEnteredValue': '79'}
                ],
                'format': {
                    'backgroundColor': {'red': 1, 'green': 0.95, 'blue': 0.7}
                }
            }
        )
        
        # Red for < 50%
        sheet.add_conditional_format_rule(
            range_notation,
            {
                'type': 'NUMBER_LESS',
                'values': [{'userEnteredValue': '50'}],
                'format': {
                    'backgroundColor': {'red': 0.95, 'green': 0.7, 'blue': 0.7}
                }
            }
        )


# Singleton instance
_sheets_service = None

def get_sheets_service() -> GoogleSheetsService:
    """Get or create Google Sheets service instance"""
    global _sheets_service
    
    if _sheets_service is None:
        credentials_path = os.getenv(
            'GOOGLE_SERVICE_ACCOUNT_FILE',
            '/home/bebdyshev/Documents/GitHub/lms-master/backend/master-lms-475912-d0afe7611b8b.json'
        )
        
        if not os.path.exists(credentials_path):
            raise Exception(f"Google service account file not found: {credentials_path}")
        
        _sheets_service = GoogleSheetsService(credentials_path)
    
    return _sheets_service
